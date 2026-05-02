import os
import glob
import pandas as pd
from openpyxl import load_workbook

input_folder = r"D:\csv1_fixed\drowsiness\RGB"
output_folder = r"D:\csvcleaned\drowsiness\RGB"

numeric_columns = ["yaw", "pitch", "roll", "left_ear", "right_ear", "avg_ear"]

os.makedirs(output_folder, exist_ok=True)

xlsx_files = glob.glob(os.path.join(input_folder, "**", "*.xlsx"), recursive=True)

print(f"Toplam bulunan dosya: {len(xlsx_files)}")

for file_path in xlsx_files:
    file_name = os.path.basename(file_path)

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        data = list(ws.values)

        if not data or len(data) < 2:
            print(f"error: {file_name} -> empty or insufficient data")
            continue

        raw_columns = list(data[0])
        columns = []
        for col in raw_columns:
            col = str(col).strip() if col is not None else ""
            col = col.replace('"', "").replace("'", "").strip()
            columns.append(col)

        rows = data[1:]
        df = pd.DataFrame(rows, columns=columns)

        if "face_detected" not in df.columns:
            print(f"error: {file_name} -> face_detected column is missing")
            print("coloumns:", list(df.columns))
            continue

        df["face_detected"] = pd.to_numeric(df["face_detected"], errors="coerce")

        valid_idx = df.index[df["face_detected"] == 1].tolist()

        if len(valid_idx) == 0:
            print(f"error: {file_name} -> there is no row with face_detected = 1")
            continue

        first_valid = valid_idx[0]
        df = df.loc[first_valid:].reset_index(drop=True)

        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
                mean_val = df[col].mean()
                df[col] = df[col].fillna(mean_val)
            else:
                print(f"warning {file_name} -> {col} there is no such column")

        relative_path = os.path.relpath(file_path, input_folder)
        output_path = os.path.join(output_folder, relative_path)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        df.to_excel(output_path, index=False)

        print(f"processed: {file_name}")

    except Exception as e:
        print(f"error: {file_name} -> {e}")
        